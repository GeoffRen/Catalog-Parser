import sys
import re
from openpyxl import load_workbook

def main(argv):
    file = input("Enter the file name:")
    wb = load_workbook(file)
    sheet_name = input("Enter the sheet name:")
    sheet = wb.get_sheet_by_name(sheet_name)
    former_course_number_set = prune_former_courses(sheet)
    write_new_sheet(wb, sheet, former_course_number_set)
    find_prerequisites(wb)
    fit_cells(wb)
    wb.save('catalog.xlsx')

def prune_former_courses(sheet):
    former_course_number_set = set()
    pattern = 'Formerly BSCI (\d+)'
    for row in range(2, sheet.max_row + 1):
        description = sheet['I' + str(row)].value
        formerCourseNumber = re.search(pattern, description, re.I)
        if formerCourseNumber:
            former_course_number_set.add(formerCourseNumber.group(1))
    return former_course_number_set

def write_new_sheet(wb, sheet, former_course_number_set):
    wb.create_sheet('catalog')
    catalog = wb.get_sheet_by_name('catalog')
    course_number_set = set()
    catalog_row = 1
    for row in range(1, sheet.max_row + 1):
        course_number = sheet['D' + str(row)].value.strip()
        if course_number not in course_number_set and course_number not in former_course_number_set:
            course_number_set.add(course_number)
            for letter in 'ABCDEFGHI':
                catalog[letter + str(catalog_row)].value = sheet[letter + str(row)].value
            catalog_row += 1

def find_prerequisites(wb):
    catalog = wb.get_sheet_by_name('catalog')
    catalog['J1'] = 'Prerequisite(s)'
    pattern = 'Prerequisite(?: or corequisite)?:? (.+?)(?:\.|$)'
    for row in range(2, catalog.max_row + 1):
        description = catalog['I' + str(row)].value
        prerequisite = re.search(pattern, description, re.I)
        if prerequisite:
            catalog['J' + str(row)].value = prerequisite.group(1)

def fit_cells(wb):
    catalog = wb.get_sheet_by_name('catalog')
    for col in catalog.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        catalog.column_dimensions[column].width = adjusted_width

if __name__ == "__main__":
    main(sys.argv)